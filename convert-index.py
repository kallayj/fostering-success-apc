import openpyxl
from openpyxl import load_workbook

# Load the Excel file
wb = load_workbook('conservation.xlsx', data_only=False)
sheet = wb.active

# Create markdown table
markdown = "| " + " | ".join(cell.value for cell in next(sheet.iter_rows())) + " |\n"
markdown += "| " + " | ".join(["---"] * sheet.max_column) + " |\n"

for row in list(sheet.iter_rows())[1:]:
    row_text = []
    for cell in row:
        if cell.hyperlink:
            row_text.append(f"[{cell.value}]({cell.hyperlink.target})")
        else:
            row_text.append(str(cell.value) if cell.value is not None else "")
    markdown += "| " + " | ".join(row_text) + " |\n"

with open('output.md', 'w') as f:
    f.write(markdown)